import openpyxl

# 需要读取的文件
book_name_xls = '/home/stealhao/下载/公安部/公交站点、线路/2019年五大公司配车生产计划表（七月）.xlsx'

sheet_name_xlsx = '7月'

# 需要写出的文件
product_name_xls = '/home/stealhao/下载/公安部/静态数据/公交/公交线路信息采集.xlsx'

# 读取excel
def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    b = []
    k = 1
    for row in sheet.rows:
        print("当前读取第",k, "行")
        if k > 5:
            a = []
            for cell in row:
                # print("当前读取第",k,"行，第",cell.column_letter ,"列")
                # print("当前读取第",k,"行，第",cell.coordinate ,"列")
                c = build_cell_column(k)
                if cell.coordinate  in c:
                    d = cell.value
                    a.append(d)
            b.append(a)
        k += 1
    write_excel_xlsx(product_name_xls,sheet_name_xlsx,b)

def build_cell_column(index):
    str_index = str(index)
    a = ['B' + str_index,'EZ' + str_index,'FA' + str_index,'FB' + str_index,'FC' + str_index,'FD' + str_index,'FE' + str_index]
    return a

# 写入excel
def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


read_excel_xlsx(book_name_xls,sheet_name_xlsx)